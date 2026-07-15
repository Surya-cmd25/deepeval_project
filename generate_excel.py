# generate_excel.py
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Konfigurasi lokasi input dan output file
INPUT_FILENAME = "hasil_evaluasi_1/hasil_run_5.json"
OUTPUT_EXCEL = "hasil_evaluasi_1/rekap_evaluasi_run_5.xlsx"

def generate_excel_report():
    if not os.path.exists(INPUT_FILENAME):
        print(f"❌ File {INPUT_FILENAME} tidak ditemukan!")
        return

    # Memastikan folder output 'hasil_evaluasi' sudah terbentuk
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

    with open(INPUT_FILENAME, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    detail_list = data.get("detail_perbandingan", [])
    
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # TAB 1: SUMMARY DASHBOARD
    # -------------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary Dashboard"
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Desain Palet Warna Profesional
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    soft_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    soft_red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F497D")
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Header Judul
    ws_sum["B2"] = "LAPORAN EVALUASI PERFORMA LLM (BASELINE)"
    ws_sum["B2"].font = title_font
    ws_sum["B3"] = f"Nama Pengujian: {data.get('nama_pengujian', 'Skenario 1')}"
    ws_sum["B3"].font = regular_font
    ws_sum["B4"] = f"Terakhir Diperbarui: {data.get('terakhir_diperbarui', '')}"
    ws_sum["B4"].font = regular_font
    
    ws_sum["B6"] = "Ringkasan Statistik Utama"
    ws_sum["B6"].font = section_font
    
    headers_sum = ["Metrik Pengujian", "Rata-rata / Total Nilai"]
    for col_idx, header in enumerate(headers_sum, start=2):
        cell = ws_sum.cell(row=7, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    total_rows = len(detail_list)
    # Catatan: Letak huruf kolom pada rumus disesuaikan karena ada penambahan kolom teks (Akurasi bergeser ke G)
    summary_metrics = [
        ("Total Kasus Uji", total_rows),
        ("Total Passed (>= 0.85)", f"=COUNTIF('Detail Perbandingan'!O9:O{total_rows+8}, \"PASSED\")"),
        ("Total Failed (< 0.85)", f"=COUNTIF('Detail Perbandingan'!O9:O{total_rows+8}, \"FAILED\")"),
        ("Rata-rata Accuracy (%)", f"=AVERAGE('Detail Perbandingan'!G9:G{total_rows+8})"),
        ("Rata-rata Response Time (s)", f"=AVERAGE('Detail Perbandingan'!H9:H{total_rows+8})"),
        ("Rata-rata G-Eval (1-5)", f"=AVERAGE('Detail Perbandingan'!I9:I{total_rows+8})"),
        ("Rata-rata AR Score", f"=AVERAGE('Detail Perbandingan'!J9:J{total_rows+8})"),
        ("Rata-rata Fact-Checking (%)", f"=AVERAGE('Detail Perbandingan'!K9:K{total_rows+8})"),
        ("Rata-rata MAE", f"=AVERAGE('Detail Perbandingan'!L9:L{total_rows+8})"),
        ("Rata-rata RMSE", f"=AVERAGE('Detail Perbandingan'!M9:M{total_rows+8})"),
        ("Rata-rata MAPE (%)", f"=AVERAGE('Detail Perbandingan'!N9:N{total_rows+8})")
    ]
    
    for row_offset, (label, val) in enumerate(summary_metrics):
        r = 8 + row_offset
        c_label = ws_sum.cell(row=r, column=2, value=label)
        c_val = ws_sum.cell(row=r, column=3, value=val)
        
        c_label.font = regular_font
        c_label.border = thin_border
        c_val.font = bold_font
        c_val.border = thin_border
        
        if "Passed" in label: c_val.fill = soft_green_fill
        elif "Failed" in label: c_val.fill = soft_red_fill
            
        if "Accuracy" in label or "Fact-Checking" in label or "MAPE" in label:
            c_val.number_format = '0.00"%"'
        elif "Time" in label or "G-Eval" in label or "AR" in label or "MAE" in label or "RMSE" in label:
            c_val.number_format = '0.00'
            
    # -------------------------------------------------------------
    # TAB 2: DETAIL PERBANDINGAN
    # -------------------------------------------------------------
    ws_det = wb.create_sheet(title="Detail Perbandingan")
    ws_det.views.sheetView[0].showGridLines = True
    ws_det.freeze_panes = "A9"  # Kunci header agar tidak bergeser
    
    ws_det["A2"] = "DATA DETAIL HASIL EVALUASI PER NOMOR SOAL"
    ws_det["A2"].font = title_font
    
    # STRUKTUR URUTAN KOLOM BARU (Teks + Seluruh Metrik Evaluasi)
    headers_det = [
        "No Soal", "Tipe Query", "Model LLM", "Query", 
        "Expected Output", "Hasil (Actual Output)", "Pipeline Accuracy (%)", 
        "Response Time (s)", "G-Eval (1-5)", "AR Score", "Fact-Checking (%)", 
        "MAE", "RMSE", "MAPE (%)", "Status Evaluasi", "Skor Aktual (E5)"
    ]
    
    for col_idx, header in enumerate(headers_det, start=1):
        cell = ws_det.cell(row=8, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for i, item in enumerate(detail_list):
        r = 9 + i
        eval_obj = item.get("evaluasi", {})
        
        # 1. Identitas & Konten Teks (Kolom A - F)
        ws_det.cell(row=r, column=1, value=item.get("no_soal")).alignment = Alignment(horizontal="center", vertical="top")
        ws_det.cell(row=r, column=2, value=item.get("query_type")).alignment = Alignment(horizontal="left", vertical="top")
        ws_det.cell(row=r, column=3, value=item.get("llm_model")).alignment = Alignment(horizontal="left", vertical="top")
        
        # Mengaktifkan wrap_text untuk kolom teks panjang agar rapi ke bawah
        cell_q = ws_det.cell(row=r, column=4, value=item.get("query"))
        cell_q.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        cell_e = ws_det.cell(row=r, column=5, value=item.get("expected_output"))
        cell_e.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        cell_a = ws_det.cell(row=r, column=6, value=item.get("actual_output"))
        cell_a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # 2. Nilai Metrik Evaluasi (Kolom G - N)
        ws_det.cell(row=r, column=7, value=item.get("pipeline_accuracy")).number_format = '0.00'
        ws_det.cell(row=r, column=8, value=item.get("response_time")).number_format = '0.00'
        ws_det.cell(row=r, column=9, value=item.get("g_eval")).number_format = '0.00'
        ws_det.cell(row=r, column=10, value=item.get("ar_score")).number_format = '0.00'
        ws_det.cell(row=r, column=11, value=item.get("fact_checking")).number_format = '0.00'
        ws_det.cell(row=r, column=12, value=item.get("mae")).number_format = '0.0000'
        ws_det.cell(row=r, column=13, value=item.get("rmse")).number_format = '0.0000'
        ws_det.cell(row=r, column=14, value=item.get("mape")).number_format = '0.00'
        
        # 3. Status Lolos & Skor Aktual (Kolom O, P)
        status_val = eval_obj.get("status", "FAILED")
        c_status = ws_det.cell(row=r, column=15, value=status_val)
        c_status.alignment = Alignment(horizontal="center", vertical="top")
        c_status.font = bold_font
        c_status.fill = soft_green_fill if status_val == "PASSED" else soft_red_fill
            
        ws_det.cell(row=r, column=16, value=eval_obj.get("skor_aktual")).number_format = '0.0000'
        
        # Menyelaraskan alignment vertikal untuk kolom metrik
        for c in range(7, 17):
            if c != 15:
                ws_det.cell(row=r, column=c).alignment = Alignment(horizontal="right", vertical="top")
        
        # Terapkan border tipis ke seluruh cell
        for c in range(1, 17):
            cell = ws_det.cell(row=r, column=c)
            if c != 15: cell.font = regular_font
            cell.border = thin_border
            
    # Auto-adjust lebar kolom otomatis (khusus untuk kolom metrik pendek)
    for ws in [ws_sum, ws_det]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if ws == ws_det and col_letter in ['D', 'E', 'F']:
                continue # Skip auto-width untuk kolom teks panjang agar tidak over-stretch
            max_len = 0
            for cell in col:
                if cell.value:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # MENENTUKAN LEBAR KHUSUS PROPORSIONAL UNTUK KOLOM TEKS PANJANG (Aura Layout Rapi)
    ws_det.column_dimensions['A'].width = 10   # No Soal
    ws_det.column_dimensions['B'].width = 14   # Tipe Query
    ws_det.column_dimensions['C'].width = 28   # Model LLM
    ws_det.column_dimensions['D'].width = 35   # Query (Pertanyaan)
    ws_det.column_dimensions['E'].width = 40   # Expected Output (Kunci)
    ws_det.column_dimensions['F'].width = 45   # Actual Output (Hasil AI)
    
    wb.save(OUTPUT_EXCEL)
    print(f"✅ Excel Laporan Lengkap Sukses Diperbarui di: '{OUTPUT_EXCEL}'")

if __name__ == "__main__":
    generate_excel_report()